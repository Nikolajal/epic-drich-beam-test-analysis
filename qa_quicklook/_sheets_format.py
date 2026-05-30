"""Formatting spec + xlsx writer for the cross-shifter sync worksheets.

Why a separate module
---------------------
Two output targets share the same look: the local ``.xlsx`` preview
operators generate before flipping ``[sheets_sync] enabled = true``,
and (next iteration) the live Google Sheet's formatting pass on the
post-push step.  Keeping the spec as plain data and the per-target
renderer as a thin adapter means a colour-palette tweak doesn't have
to be edited twice.

Spec contents are intentionally narrow:

  - **Column widths** by name — operators read the run-id, notes, and
    radiators columns most, so those get more room than the
    autoscaled default.
  - **Header style** — bold + dark fill + white text, frozen so the
    column names stay visible while scrolling 260 runs.
  - **Quality / state column shading** — values mapped through a tiny
    palette function, light fills so the text on top stays readable.
  - **Zebra rows** for the long sheets (``runs``, ``audit``).

What's NOT in here
------------------
Anything that needs evaluating dynamically (conditional formatting
rules, sparklines, charts).  Those belong in a v2 if/when operators
ask for them — the cell-level look is the first-order win.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


#  Pattern matches Phase B worksheet titles like ``Runs (2025)`` /
#  ``Radiators`` → base name ``runs`` / ``radiators``.  Lower-cases
#  the result so the FORMATS / TAB_COLORS dicts stay keyed by a
#  small, stable lowercase set even though the on-Sheet titles are
#  capitalised.
_TITLE_RE = re.compile(r"^([A-Za-z_]+)(?:\s*\([^)]+\))?\s*$")


def base_name(title: str) -> str:
    """Strip a year suffix off a worksheet title and lower-case it.

    ``"Runs (2025)"`` → ``"runs"``; ``"Audit"`` → ``"audit"``.
    Unknown shapes (no match) return the title verbatim (lowered)
    so a missing FORMATS entry surfaces as "no formatting" rather
    than a crash.
    """
    m = _TITLE_RE.match(title.strip())
    return (m.group(1) if m else title.strip()).lower()


# ---------------------------------------------------------------------------
# Palette — kept in one place so both targets stay visually consistent
# and so a "we want a different shade of green" comes down to one line.
# Hex strings WITHOUT the leading "#" so they slot into openpyxl
# PatternFill (8-digit "AARRGGBB" expected) and the Sheets API
# colorRgb representation (R, G, B as 0..1 floats) with a tiny helper.
# ---------------------------------------------------------------------------


#  Operator-supplied palette (2026-05-29).  Five colours; everything
#  else is derived (lightened tints for body fills).  Documented at
#  https://coolors.co/ff6b6b-0bda51-afdbf5-9b8e8e-2a2929.
PAL_NEAR_BLACK = "2A2929"
PAL_GREEN      = "0BDA51"
PAL_RED        = "FF6B6B"
PAL_BLUE       = "AFDBF5"
PAL_WARM_GRAY  = "9B8E8E"

#  Derived role colours — mirror the dashboard's ``theme.light_palette``
#  so the Sheet and the operator's desktop QA tool track each other
#  visually.  See ``qa_quicklook/theme.py`` for the source of truth.
PAL_WARNING_AMBER = "E0A40A"   # dashboard light_palette().warning
PAL_TEXT_MUTED    = "6B6968"   # dashboard light_palette().text_muted
PAL_BG            = "FAFAFB"   # dashboard light_palette().bg
PAL_SURFACE_ALT   = "F1F3F6"   # dashboard light_palette().surface_alt

HEADER_BG_HEX = PAL_NEAR_BLACK
HEADER_FG_HEX = "FFFFFF"
#  Zebra + meta-key tints match the dashboard's surface_alt + bg —
#  cell text stays legible against either (no alpha in xlsx; openpyxl
#  can't render transparency).
ZEBRA_HEX     = PAL_SURFACE_ALT
META_KEY_HEX  = PAL_BG

#  Quality semantic colours: green / amber / red / blue.
#    good     — the dashboard's success
#    warning  — the dashboard's warning amber
#    bad      — the dashboard's danger red
#    need QA  — the dashboard's info sky; operator hasn't judged yet
#               (auto-set by the writer pipeline before review).
#               Sourced from production: 213/1001 runs carry this
#               value as of 2026-05-29; the dropdown must accept it.
QUALITY_PALETTE: dict[str, str] = {
    "good":    PAL_GREEN,
    "warning": PAL_WARNING_AMBER,
    "bad":     PAL_RED,
    "need QA": PAL_BLUE,
    #  Phase D.12: ``test`` isn't a verdict — it's an operator-marked
    #  category for runs taken without beam / for calibration only.
    #  Warm grey distinguishes them visually from the four-way quality
    #  flag without implying judgment.
    "test":    PAL_WARM_GRAY,
}

#  Beam ON/OFF — the strictest binary on the Sheet.  Uppercase keys
#  because Phase D.2 normalises display values to ``ON``/``OFF`` even
#  though the database stores lowercase.
BEAM_PALETTE: dict[str, str] = {
    "ON":  PAL_GREEN,
    "OFF": PAL_RED,
}

#  Polarity Pos/Neg — same chip treatment as Beam (strict dropdown,
#  display abbreviated).  Uses sky blue + amber instead of green/red
#  so the column is visually distinct from Beam and so neither value
#  carries an implicit good/bad signal — physics charge, not a verdict.
POLARITY_PALETTE: dict[str, str] = {
    "Pos": PAL_BLUE,
    "Neg": PAL_WARNING_AMBER,
}



#  Job states track the same green/red poles for healthy / failed.
#  Killed + abandoned fall to muted warm-gray (palette's mute slot).
JOB_STATE_PALETTE: dict[str, str] = {
    "running":   PAL_GREEN,
    "success":   PAL_BLUE,
    "error":     PAL_RED,
    "killed":    PAL_WARM_GRAY,
    "abandoned": PAL_WARM_GRAY,
}

#  Audit sources — only ``sheet`` and ``dashboard`` get strong
#  colours (the two interactive paths shifters care about); writer
#  sources fall to muted warm-gray so audit-log scrolling stays calm.
AUDIT_SOURCE_PALETTE: dict[str, str] = {
    "dashboard":   PAL_BLUE,
    "sheet":       PAL_GREEN,
    "legacy":      PAL_WARM_GRAY,
    "lightdata":   PAL_WARM_GRAY,
    "recodata":    PAL_WARM_GRAY,
    "recotrack":   PAL_WARM_GRAY,
    "calibration": PAL_WARM_GRAY,
}


# ---------------------------------------------------------------------------
# Spec
# ---------------------------------------------------------------------------


@dataclass
class SheetFormat:
    """Per-worksheet declarative formatting spec.

    ``column_widths`` maps a column NAME (the header value, not an
    index) to its width in xlsx character units / Sheets pixels.  Any
    column not in the map defaults to ``default_width``.

    Color columns are by name too — the renderer looks the column up
    in the header row and applies the palette to every body cell.

    ``beam_column`` gets the strictest treatment — a hard data-
    validation dropdown limited to the BEAM_PALETTE keys, with
    green/red conditional formatting per the dashboard's parity.
    """

    name: str
    frozen_rows: int = 1
    frozen_cols: int = 0      # Phase D.9 — pinned identity columns
    default_width: int = 14
    column_widths: dict[str, int] = field(default_factory=dict)
    numeric_formats: dict[str, str] = field(default_factory=dict)  # display name → number pattern
    #  Phase D.11: per-column alignment overrides for the long-text
    #  columns that read better left-aligned than centered (today: Notes).
    left_aligned_columns: list[str] = field(default_factory=list)
    zebra_rows: bool = False
    quality_column: str = ""
    state_column: str = ""
    source_column: str = ""
    beam_column: str = ""
    polarity_column: str = ""
    bold_first_column: bool = False


#  Header row index — Phase B added a row-0 banner ("Last push: …")
#  so the column-name header is row 1 (0-indexed).  Every worksheet
#  except runlists uses ``frozen_rows = 2`` so both banner + header
#  stay visible while scrolling.
#  Runlists is vertical (name / campaign / count metadata rows above
#  the run-id block) and uses ``frozen_rows = 4``.

#  FORMATS column-width keys use the operator-facing display names
#  (Phase D.1) since the rendered header row carries those.  The
#  storage-key → display-name mapping lives in
#  ``sheets_sync.FIELD_DISPLAY``.
FORMATS: dict[str, SheetFormat] = {
    "runs": SheetFormat(
        name="runs",
        frozen_rows=3,           # banner + group band + column header
        frozen_cols=3,           # Run ID + N spills + Quality (Phase D.9)
        zebra_rows=True,
        quality_column="Quality",
        beam_column="Status",
        polarity_column="Polarity",
        numeric_formats={
            "Temp (°C)":   "0.0",     # X.X — one decimal, trailing zero preserved
            "V_bias (V)":  "0.00",    # X.XX — two decimals
        },
        left_aligned_columns=["Notes"],   # long free-form text reads left
        column_widths={
            "Run ID":             18,
            "Mirror A (mm)":      12,   # aerogel mirror (Phase D.8 short form)
            "Mirror G (mm)":      12,   # gas mirror (Phase D.8 short form)
            "Mirror aerogel (mm)": 16,   # legacy fallback
            "Mirror (mm)":        12,   # legacy fallback
            "Mirror gas (mm)":    14,   # legacy fallback
            "Status":             10,   # display name for beam_status (Phase D.7)
            "Particle":           10,
            "Polarity":           10,
            "Collimators":        12,
            "Energy (GeV)":       12,
            "Trigger":            22,
            "RDO firmware":       18,
            "Timing firmware":    32,
            "Temp (°C)":          10,   # ~-100…+50°C max range + header
            "Temperature (°C)":   16,   # legacy key — harmless fallback
            "V_bias (V)":         11,
            "Calibration run":    20,
            "DCR scan run":       20,
            "N spills":          10,    # Run group, frozen — keep readable
            "Radiator IDs":       14,
            "GEM":                10,
            "ALTAI":              10,   # ALTAI tracking-set run number
            "Triggers":           18,
            "Quality":            11,
            "Notes":             120,   # double the previous 60 per 2026-05-29 design call
            "Radiator description": 32,
            "Radiator gas":       18,
            "DThr":               6,    # Δ threshold — 1-2 digit values; tight header
            "Op":                 5,    # Op mode — single digit values
            "Δ threshold":        12,   # legacy fallback
            "Op mode":            10,   # legacy fallback
        },
    ),
    "runlists": SheetFormat(
        name="runlists",
        frozen_rows=5,           # banner + name + year + campaign + count
        zebra_rows=False,        # banding clashes with the meta rows
        bold_first_column=True,  # column A holds row labels
        default_width=20,        # each runlist column carries run ids (timestamp width)
        column_widths={
            "Name":     14,   # column A row labels
        },
    ),
    "radiators": SheetFormat(
        name="radiators",
        frozen_rows=2,
        zebra_rows=True,
        left_aligned_columns=["Runs ranges"],   # long; reads left
        column_widths={
            "ID":               6,
            "Type":             12,
            "Refractive index": 16,
            "Tag":              14,
            "Depth (cm)":       11,
            "Side":             8,
            "N runs":           8,
            "Runs ranges":      80,   # can hold several "first..last" entries
            "First run":        18,   # legacy fallback
            "Last run":         18,   # legacy fallback
        },
    ),
    "audit": SheetFormat(
        name="audit",
        frozen_rows=2,
        zebra_rows=True,
        source_column="Source",
        column_widths={
            "At":        20,
            "Source":    14,
            "Run":       18,
            "Field":     16,
            "Old value": 30,
            "New value": 30,
        },
    ),
}


# ---------------------------------------------------------------------------
# xlsx writer
# ---------------------------------------------------------------------------


class _OpenpyxlMissing(ImportError):
    """Raised when ``openpyxl`` isn't installed.

    The xlsx writer is optional dependency — operators who only push
    to Google Sheets directly don't need the package.  Caller (the
    CLI) catches and prints a copy-pasteable install hint.
    """


def to_xlsx(
    rendered: dict[str, list[list[Any]]],
    path: Path,
    *,
    formats: dict[str, SheetFormat] = FORMATS,
) -> Path:
    """Write the rendered worksheets to an ``.xlsx`` file.

    Returns the resolved output path.  The file is openable in Excel,
    LibreOffice, Numbers, and Google Sheets (via File → Import) —
    operators preview the look locally before flipping the
    ``enabled = true`` switch in ``qa_quicklook.toml``.

    Raises ``_OpenpyxlMissing`` when the package isn't installed,
    with a copy-pasteable install hint.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise _OpenpyxlMissing(
            "openpyxl is required for --to-xlsx — run "
            "`.venv/bin/pip install openpyxl`.  "
            f"(original error: {exc})"
        ) from exc

    wb = Workbook()
    #  Workbook ships a default "Sheet" we don't want — drop it
    #  before adding ours so the file opens on the runs sheet.
    if wb.sheetnames:
        wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor=HEADER_BG_HEX)
    header_font = Font(bold=True, color=HEADER_FG_HEX)
    zebra_fill = PatternFill("solid", fgColor=ZEBRA_HEX)
    meta_key_fill = PatternFill("solid", fgColor=META_KEY_HEX)
    centered = Alignment(vertical="center")

    for name, rows in rendered.items():
        ws = wb.create_sheet(name)
        fmt = formats.get(base_name(name), SheetFormat(name=name))
        if not rows:
            #  Empty worksheet — still create + style so the tab
            #  shows up rather than vanishing.
            ws["A1"] = "(no data)"
            ws["A1"].font = Font(italic=True, color="6B7280")
            continue

        #  Layout: row 1 = banner, row 2 = column header, rows 3+ = data.
        #  Phase D.5 ``runs`` worksheets have a group band on row 2 +
        #  column header on row 3.  ``runlists`` keeps the banner +
        #  metadata-stack layout (banner / name / year / campaign / count).
        #  ``fmt.frozen_rows`` is the count of rows to freeze;
        #  ``header_row_1idx`` is the 1-indexed row carrying column names.
        bn = base_name(name)
        group_row_1idx: int | None = None
        if bn == "runs" and len(rows) >= 3:
            group_row_1idx = 2
            header_row_1idx = 3
        else:
            header_row_1idx = 2
        header = [str(c) for c in rows[header_row_1idx - 1]] if len(rows) >= header_row_1idx else []
        ncols = len(header) if header else (len(rows[0]) if rows else 0)

        #  Write all cells first; styling next so column-letter
        #  lookups have something to point at.
        for r, row in enumerate(rows, start=1):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=_xlsx_value(val))

        #  Banner row (row 1) — italic + muted so it reads as a
        #  caption, not a header.
        if rows:
            banner_cell = ws.cell(row=1, column=1)
            banner_cell.font = Font(italic=True, color="6B7280", size=11)

        #  Group-band row (runs only) — neutral tint, bold, centered,
        #  merged per span.
        if group_row_1idx is not None and len(rows) >= group_row_1idx:
            from openpyxl.utils import get_column_letter as _gcl
            group_row = [str(c) for c in rows[group_row_1idx - 1]]
            group_fill = PatternFill("solid", fgColor=_GROUP_BAND_BG)
            for start, end, _label in _group_spans(group_row):
                #  Style first then merge — merging clears style on
                #  continuation cells which we don't need anyway.
                for c in range(start + 1, end + 1):
                    cell = ws.cell(row=group_row_1idx, column=c)
                    cell.fill = group_fill
                    cell.font = Font(bold=True, size=11)
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center",
                    )
                if end - start > 1:
                    ws.merge_cells(
                        start_row=group_row_1idx, end_row=group_row_1idx,
                        start_column=start + 1, end_column=end,
                    )

        #  Column header row — bold, dark fill, white text, centered.
        if len(rows) >= header_row_1idx:
            for c in range(1, ncols + 1):
                cell = ws.cell(row=header_row_1idx, column=c)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = centered

        #  Column widths.  Honour the per-column spec; fall back to
        #  the worksheet's default_width.  Capped at 80 chars so a
        #  pathological long row doesn't explode the layout.
        for c in range(1, ncols + 1):
            col_name = header[c - 1] if c - 1 < len(header) else ""
            w = fmt.column_widths.get(col_name, fmt.default_width)
            ws.column_dimensions[get_column_letter(c)].width = min(w, _MAX_COLUMN_WIDTH)

        #  Freeze the configured rows + columns (Phase D.9).  openpyxl's
        #  ``freeze_panes`` takes the top-left cell of the unfrozen
        #  region: row ``frozen_rows + 1``, column ``frozen_cols + 1``.
        if fmt.frozen_rows or fmt.frozen_cols:
            ws.freeze_panes = ws.cell(
                row=fmt.frozen_rows + 1,
                column=fmt.frozen_cols + 1,
            )

        #  Quality / state / source colour-coding.  Each path looks
        #  the column up by name, falls through silently when the
        #  column isn't present in this worksheet's header (defensive
        #  for the test path which renders synthetic data).  Data
        #  rows start at ``header_row_1idx + 1``.
        body_start = header_row_1idx + 1
        #  Phase D.8: centre every body cell so numbers + tags + text
        #  read on a single vertical axis.  Applied first; the tag-
        #  column pass below will override font (bold) without
        #  touching alignment.  Phase D.11: per-column override —
        #  the ``left_aligned_columns`` list (today: Notes) gets a
        #  LEFT alignment instead of CENTER.
        centered_body = Alignment(horizontal="center", vertical="center")
        left_body = Alignment(horizontal="left", vertical="center")
        left_aligned_idxs = {
            header.index(c) + 1 for c in fmt.left_aligned_columns
            if c in header
        }
        for r in range(body_start, len(rows) + 1):
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = left_body if c in left_aligned_idxs else centered_body
        _shade_by_palette(ws, header, rows, fmt.quality_column,
                          QUALITY_PALETTE, body_start=body_start)
        _shade_by_palette(ws, header, rows, fmt.state_column,
                          JOB_STATE_PALETTE, body_start=body_start)
        _shade_by_palette(ws, header, rows, fmt.source_column,
                          AUDIT_SOURCE_PALETTE, body_start=body_start)
        _shade_by_palette(ws, header, rows, fmt.beam_column,
                          BEAM_PALETTE, body_start=body_start)
        _shade_by_palette(ws, header, rows, fmt.polarity_column,
                          POLARITY_PALETTE, body_start=body_start)

        #  Tag-style every palette-bound column: centered + bold so the
        #  coloured fill reads as a chip rather than a wash.  Cap at the
        #  worksheet's data range so blank trailing rows aren't styled.
        for col_name in _tag_columns_for(fmt):
            if col_name not in header:
                continue
            col_idx = header.index(col_name) + 1
            for r in range(body_start, len(rows) + 1):
                cell = ws.cell(row=r, column=col_idx)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center",
                )
                cell.font = Font(bold=True)

        #  Zebra rows — only paint cells that don't already carry a
        #  semantic fill so the quality column stays its own colour.
        if fmt.zebra_rows:
            for r in range(body_start, len(rows) + 1):
                if (r - body_start) % 2 == 0:
                    continue   # leave first/odd rows white
                for c in range(1, ncols + 1):
                    cell = ws.cell(row=r, column=c)
                    if _cell_has_fill(cell):
                        continue
                    cell.fill = zebra_fill

        if fmt.bold_first_column:
            #  Bold + tinted-fill the first column from row 1 onwards;
            #  the column header row (row 2) already carries the dark
            #  header style so we skip it explicitly.
            for r in range(1, len(rows) + 1):
                if r == header_row_1idx:
                    continue
                cell = ws.cell(row=r, column=1)
                cell.font = Font(bold=True)
                if r != 1:   # don't tint the banner row
                    cell.fill = meta_key_fill

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path


def _xlsx_value(v: Any) -> Any:
    """Coerce a rendered cell into something openpyxl can write.

    Lists/dicts that snuck past the JSON-encoding pass at render
    time (defensive — shouldn't happen) get stringified rather than
    crashing the workbook save.
    """
    if v is None:
        return ""
    if isinstance(v, (list, dict)):
        return str(v)
    return v


def _shade_by_palette(
    ws,
    header: list[str],
    rows: list[list[Any]],
    column_name: str,
    palette: dict[str, str],
    *,
    body_start: int = 2,
) -> None:
    """Fill every body cell in ``column_name`` with the palette colour.

    The cell ``Quality`` of value ``"warning"`` gets the amber fill,
    ``"good"`` the green, etc.  Combined with the bold + centered
    alignment from the tag-style pass, the result is Sheets' native
    "Chip" dropdown look — coloured pill in a narrow cell.

    Key match is case-insensitive on a strip so a stray space or
    casing typo still chips (Quality's lenient dropdown allows
    operators free-form input).
    """
    if not column_name or column_name not in header:
        return
    from openpyxl.styles import PatternFill
    palette_lower = {k.lower(): v for k, v in palette.items()}
    col_idx = header.index(column_name) + 1
    for r in range(body_start, len(rows) + 1):
        cell = ws.cell(row=r, column=col_idx)
        key = str(cell.value or "").strip().lower()
        hex_ = palette_lower.get(key)
        if hex_:
            cell.fill = PatternFill("solid", fgColor=hex_)


def _cell_has_fill(cell) -> bool:
    """True iff a cell already carries a non-default PatternFill.

    Defensive against openpyxl's "no fill" representation drifting
    between versions — the rgb attribute is sometimes ``"00000000"``
    (transparent) and sometimes literally ``None``.
    """
    if cell.fill is None:
        return False
    fg = cell.fill.fgColor
    if fg is None:
        return False
    rgb = getattr(fg, "rgb", None)
    return bool(rgb) and rgb != "00000000"


# ---------------------------------------------------------------------------
# Live Sheet — translate the same FORMATS spec into Sheets API
# batchUpdate JSON.  Shared with the xlsx writer so a palette tweak
# lands once and shows up in both targets.
#
# Strategy: prefer Sheets-native primitives (conditional formatting
# rules, banded ranges, data validation) over per-cell repaints.  The
# rules survive operator edits to the cells — we don't have to re-
# paint anything when a shifter changes ``quality`` from "good" to
# "bad", Sheets just re-evaluates the rule.
# ---------------------------------------------------------------------------


#  Soft tab tints so each worksheet is eyeball-findable at the bottom
#  of the Sheet UI.  Muted enough that the active-tab outline still
#  reads as the dominant signal.
#  Tab colours keyed by base name (use ``base_name("runs (2025)")`` →
#  ``"runs"`` so all year-suffixed tabs share their family colour).
#  ``qa_results`` gets the most distinct hue (purple) to underline
#  the "derived from QA / processing — not data-taking metadata"
#  provenance.
#  Tab colours keyed by base name (lower-cased via ``base_name``).
#  Each tab gets one of the palette members so the bottom strip
#  reads as a colour-coded index of the whole workbook.
TAB_COLORS: dict[str, str] = {
    "runs":      PAL_GREEN,        # the main data-taking tab
    "runlists":  PAL_BLUE,         # selections / analysis intents
    "radiators": PAL_WARM_GRAY,    # auxiliary catalog
    "audit":     PAL_RED,          # edit history — drawing attention
}

#  Sheets defaults a column to ~100 px.  We translate the xlsx
#  character-width units (~7 px/char + margin) so the live Sheet and
#  the local xlsx track each other's widths within rounding.
_PX_PER_CHAR = 8


def _hex_to_rgb(hex6: str) -> dict:
    """Convert a 6-digit hex string to Sheets API ``Color`` object.

    Sheets wants ``{red, green, blue}`` as 0..1 floats — palette
    entries are 6-digit hex for human readability, this is the bridge.
    """
    h = hex6.lstrip("#")
    return {
        "red":   int(h[0:2], 16) / 255.0,
        "green": int(h[2:4], 16) / 255.0,
        "blue":  int(h[4:6], 16) / 255.0,
    }


#  Group-band background — Phase D.7 darkened from PAL_SURFACE_ALT
#  to a warmer mid-grey so the chapter labels read as a defined
#  band rather than melting into the row backgrounds.  Still in
#  the dashboard's warm-grey family — visual rhyme preserved.
_GROUP_BAND_BG = "D6D0CE"

#  Vertical divider colour for the between-group seam.  Picked to
#  match the dashboard's ``border`` role (light-mode ``#D6D9DE``)
#  but a touch darker so it reads as a chapter break rather than
#  a column boundary.
_GROUP_DIVIDER = "9B8E8E"


#  Tag-styled column width — narrow enough that the coloured fill
#  reads as a chip rather than a horizontal bar.  Picked so every
#  documented value across palettes fits without truncation:
#    Beam   → "ON"/"OFF" (≤ 3 chars)
#    Quality→ "warning"/"need QA" (7 chars, the longest)
#    Source → "calibration" (11 chars, the longest)
#  ``_TAG_COLUMN_WIDTHS`` overrides ``column_widths`` for any column
#  identified as a tag column.
_TAG_COLUMN_WIDTHS: dict[str, int] = {
    "Status":   7,   # was "Beam"; Phase D.7 renamed for group-nested cleanliness
    "Quality":  11,  # 9 was a touch cramped; the dropdown caret needs room
    "Source":   12,
    "Polarity": 8,   # Phase D.10 — narrow now that values are Pos/Neg
}


#  Hard cap on any column width so a pathological long-string column
#  doesn't push the data tab off-screen.  Picked to give Notes (120
#  chars by spec) full effect while still bounding the worst case.
_MAX_COLUMN_WIDTH = 150


def _tag_columns_for(fmt: SheetFormat) -> list[str]:
    """Which display-name columns get tag styling on this worksheet.

    Pulls from the palette-bound slots on ``SheetFormat`` so the set
    stays driven by the spec — adding a new palette column to a format
    automatically picks up the tag treatment.
    """
    return [c for c in (
        fmt.beam_column, fmt.quality_column,
        fmt.source_column, fmt.state_column,
        fmt.polarity_column,
    ) if c]


def _group_spans(group_row: list[str]) -> list[tuple[int, int, str]]:
    """Walk a group-row (one label-or-empty per column) into spans.

    Each span is ``(start_col_idx, end_col_idx_exclusive, group_name)``.
    The renderer only puts a label on the FIRST cell of each group,
    so any non-empty cell starts a new span and runs until the next
    non-empty cell (or end of row).  Ungrouped leading columns (like
    ``Run ID``) get no span.
    """
    spans: list[tuple[int, int, str]] = []
    i = 0
    while i < len(group_row):
        label = str(group_row[i]).strip()
        if not label:
            i += 1
            continue
        j = i + 1
        while j < len(group_row) and not str(group_row[j]).strip():
            j += 1
        spans.append((i, j, label))
        i = j
    return spans


def sheets_format_requests(
    rendered: dict[str, list[list[Any]]],
    sheet_id_by_title: dict[str, int],
    existing_bandings_by_sheet: dict[int, list[int]] | None = None,
    existing_conditional_rules_by_sheet: dict[int, int] | None = None,
    existing_merges_by_sheet: dict[int, list[dict]] | None = None,
    *,
    formats: dict[str, SheetFormat] = FORMATS,
) -> list[dict]:
    """Build the ``spreadsheets.batchUpdate`` request list for formatting.

    ``sheet_id_by_title`` maps worksheet name → numeric ``sheetId``
    (the gid the Sheets API requires; not the same thing as the
    spreadsheet id).  ``existing_bandings_by_sheet`` /
    ``existing_conditional_rules_by_sheet`` (both optional, both
    keyed by sheetId) carry per-sheet counts so this function can
    emit ``deleteBanding`` / ``deleteConditionalFormatRule`` requests
    first — required because addBanding + addConditionalFormatRule
    aren't idempotent on their own (re-calling them duplicates the
    rules).

    The returned list is in batchUpdate order: deletes first, then
    structural updates, then formatting adds.  Caller wraps it in
    ``{"requests": <list>}`` and posts to the API.
    """
    existing_bandings_by_sheet = existing_bandings_by_sheet or {}
    existing_conditional_rules_by_sheet = existing_conditional_rules_by_sheet or {}
    existing_merges_by_sheet = existing_merges_by_sheet or {}
    requests: list[dict] = []

    #  --- Phase 1: delete existing dynamic-format objects ---
    #  Bandings: each banding_id can only be deleted once; collect
    #  every id we've ever heard about and clear them.
    for sheet_id, banding_ids in existing_bandings_by_sheet.items():
        for bid in banding_ids:
            requests.append({"deleteBanding": {"bandedRangeId": bid}})

    #  Conditional rules: deleted by (sheetId, index) — index 0 first,
    #  but since each delete shifts the remaining indices DOWN, we
    #  delete from the top each time and that takes care of the shift.
    for sheet_id, count in existing_conditional_rules_by_sheet.items():
        for _ in range(count):
            requests.append({
                "deleteConditionalFormatRule": {
                    "sheetId": sheet_id, "index": 0,
                }
            })

    #  Merges: ``mergeCells`` is non-idempotent (re-merging the same
    #  range raises 400 "already merged") so we delete every existing
    #  merge before issuing fresh ones for the group-band row.
    for sheet_id, merges in existing_merges_by_sheet.items():
        for m in merges:
            requests.append({"unmergeCells": {"range": m}})

    #  Data validation: ``setDataValidation`` doesn't replace, it
    #  COEXISTS — leaving a stale Quality dropdown clinging to whatever
    #  column used to be Quality before a reorder.  Hit observed
    #  2026-05-29 on the ``Δ threshold`` / ``Op mode`` / ``N spills``
    #  columns after the Phase D.5 group reshuffle.  Fix: wipe
    #  validation for every worksheet we're about to format; the new
    #  per-column rules get added back below.  Sending the request
    #  without a ``rule`` field clears validation on the range.
    for title in rendered.keys():
        sheet_id = sheet_id_by_title.get(title)
        if sheet_id is None:
            continue
        requests.append({"setDataValidation": {
            "range": {"sheetId": sheet_id},  # full-sheet range
        }})

    #  --- Phase 2: structural updates per worksheet ---
    #  Tab ordering: the ``rendered`` dict's insertion order IS the
    #  desired tab strip order (newest year first per the Phase C
    #  reorder).  We assign sequential ``index`` values starting from
    #  0 so existing tabs get shuffled into place even when they were
    #  created in a different order on an earlier push.  Sheets API
    #  applies these moves in batchUpdate order — each ``index`` is
    #  the post-move position, so identical sequential values work.
    for tab_index, (title, rows) in enumerate(rendered.items()):
        sheet_id = sheet_id_by_title.get(title)
        if sheet_id is None or not rows:
            continue
        bname = base_name(title)
        fmt = formats.get(bname, SheetFormat(name=title))
        #  Layout: row 0 = banner.  Runs tabs since Phase D.5 have a
        #  group band on row 1 + column header on row 2; everyone else
        #  has the column header on row 1.  Detect by checking for the
        #  Phase D.5 group row signature (first cell empty, at least
        #  one cell carrying a known group label).
        group_row_idx: int | None = None
        header_row = 1 if len(rows) >= 2 else 0
        if bname == "runs" and len(rows) >= 3:
            #  Group row sits at index 1, column header at index 2.
            group_row_idx = 1
            header_row = 2
        header = [str(c) for c in rows[header_row]] if len(rows) >= header_row + 1 else []
        ncols = len(header) if header else (len(rows[0]) if rows else 0)
        nrows = len(rows)

        #  Tab colour + frozen rows + frozen cols + index.
        #  ``updateSheetProperties`` is idempotent on its own — fields
        #  mask scopes the merge so unrelated properties (sheet title,
        #  etc.) survive.
        tab_color = TAB_COLORS.get(bname)
        sheet_props: dict = {
            "sheetId": sheet_id,
            "index": tab_index,
            "gridProperties": {
                "frozenRowCount": fmt.frozen_rows,
                "frozenColumnCount": fmt.frozen_cols,
            },
        }
        fields = ("index,gridProperties.frozenRowCount,"
                  "gridProperties.frozenColumnCount")
        if tab_color:
            sheet_props["tabColorStyle"] = {"rgbColor": _hex_to_rgb(tab_color)}
            fields += ",tabColorStyle"
        requests.append({"updateSheetProperties": {
            "properties": sheet_props,
            "fields": fields,
        }})

        #  Column widths.  One ``updateDimensionProperties`` per
        #  column — the API accepts a range, but each column has a
        #  different desired width so a per-column request is the
        #  natural shape.  Tag columns override to the narrow
        #  ``_TAG_COLUMN_WIDTHS`` value so the coloured fill reads as
        #  a chip.
        tag_cols = set(_tag_columns_for(fmt))
        for c, col_name in enumerate(header):
            if col_name in tag_cols and col_name in _TAG_COLUMN_WIDTHS:
                w_chars = _TAG_COLUMN_WIDTHS[col_name]
            else:
                w_chars = fmt.column_widths.get(col_name, fmt.default_width)
            requests.append({"updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": c,
                    "endIndex": c + 1,
                },
                "properties": {"pixelSize": int(min(w_chars, _MAX_COLUMN_WIDTH) * _PX_PER_CHAR)},
                "fields": "pixelSize",
            }})

        #  Banner row style (row 0) — italic, muted gray, slightly
        #  smaller.  Reads as a caption above the column header.
        if nrows >= 1:
            requests.append({"repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0, "endRowIndex": 1,
                    "startColumnIndex": 0, "endColumnIndex": max(1, ncols),
                },
                "cell": {"userEnteredFormat": {
                    "backgroundColorStyle": {"rgbColor": _hex_to_rgb("FFFFFF")},
                    "textFormat": {
                        "foregroundColorStyle": {"rgbColor": _hex_to_rgb("6B7280")},
                        "italic": True,
                        "fontSize": 10,
                    },
                }},
                "fields": "userEnteredFormat(backgroundColorStyle,textFormat)",
            }})

        #  Group-band row (Phase D.5, runs only) — neutral tint, bold,
        #  centered.  Each span is merged so the label sits in the
        #  visual middle of its group.  Phase D.7 adds a vertical
        #  divider on the right edge of every span (except the last)
        #  so chapters read as discrete sections.
        if group_row_idx is not None and len(rows) > group_row_idx and ncols > 0:
            group_row = [str(c) for c in rows[group_row_idx]]
            spans = _group_spans(group_row)
            for start, end, _label in spans:
                #  Style the whole span (label cell + continuations).
                requests.append({"repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": group_row_idx,
                        "endRowIndex": group_row_idx + 1,
                        "startColumnIndex": start,
                        "endColumnIndex": end,
                    },
                    "cell": {"userEnteredFormat": {
                        "backgroundColorStyle": {"rgbColor": _hex_to_rgb(_GROUP_BAND_BG)},
                        "textFormat": {"bold": True, "fontSize": 11},
                        "horizontalAlignment": "CENTER",
                        "verticalAlignment": "MIDDLE",
                    }},
                    "fields": "userEnteredFormat(backgroundColorStyle,textFormat,horizontalAlignment,verticalAlignment)",
                }})
                if end - start > 1:
                    requests.append({"mergeCells": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": group_row_idx,
                            "endRowIndex": group_row_idx + 1,
                            "startColumnIndex": start,
                            "endColumnIndex": end,
                        },
                        "mergeType": "MERGE_ALL",
                    }})

            #  Vertical chapter dividers — right border on the last
            #  column of every group except the final one.  Applied
            #  from the group-band row down to the bottom of the
            #  data so the divider runs the full visible height.
            for i, (_start, end, _label) in enumerate(spans):
                if i == len(spans) - 1:
                    continue  # no divider after the last group
                requests.append({"updateBorders": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": group_row_idx,
                        "endRowIndex": nrows,
                        "startColumnIndex": end - 1,
                        "endColumnIndex": end,
                    },
                    "right": {
                        "style": "SOLID_MEDIUM",
                        "colorStyle": {"rgbColor": _hex_to_rgb(_GROUP_DIVIDER)},
                    },
                }})

        #  Column header row style — bold + dark fill + white text.
        if nrows >= header_row + 1 and ncols > 0:
            requests.append({"repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": header_row, "endRowIndex": header_row + 1,
                    "startColumnIndex": 0, "endColumnIndex": ncols,
                },
                "cell": {"userEnteredFormat": {
                    "backgroundColorStyle": {"rgbColor": _hex_to_rgb(HEADER_BG_HEX)},
                    "textFormat": {
                        "foregroundColorStyle": {"rgbColor": _hex_to_rgb(HEADER_FG_HEX)},
                        "bold": True,
                    },
                    "verticalAlignment": "MIDDLE",
                }},
                "fields": "userEnteredFormat(backgroundColorStyle,textFormat,verticalAlignment)",
            }})

        body_start = header_row + 1   # 0-indexed first body row

        #  Banded rows — Sheets-native zebra striping over the body.
        if fmt.zebra_rows and nrows > body_start:
            requests.append({"addBanding": {"bandedRange": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": body_start, "endRowIndex": nrows,
                    "startColumnIndex": 0, "endColumnIndex": ncols,
                },
                "rowProperties": {
                    "firstBandColorStyle": {"rgbColor": _hex_to_rgb("FFFFFF")},
                    "secondBandColorStyle": {"rgbColor": _hex_to_rgb(ZEBRA_HEX)},
                },
            }}})

        #  Bold first column on bold_first_column sheets (today:
        #  runlists).  Skip the banner row + the column-header row;
        #  paint everything else.
        if fmt.bold_first_column and nrows > body_start:
            requests.append({"repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": body_start, "endRowIndex": nrows,
                    "startColumnIndex": 0, "endColumnIndex": 1,
                },
                "cell": {"userEnteredFormat": {
                    "backgroundColorStyle": {"rgbColor": _hex_to_rgb(META_KEY_HEX)},
                    "textFormat": {"bold": True},
                }},
                "fields": "userEnteredFormat(backgroundColorStyle,textFormat)",
            }})

        #  Conditional formatting rules — quality / state / source
        #  columns.  One rule per palette key so the colour follows
        #  whatever value the cell holds, even after operator edits.
        for col_name, palette in (
            (fmt.quality_column,  QUALITY_PALETTE),
            (fmt.state_column,    JOB_STATE_PALETTE),
            (fmt.source_column,   AUDIT_SOURCE_PALETTE),
            (fmt.beam_column,     BEAM_PALETTE),
            (fmt.polarity_column, POLARITY_PALETTE),
        ):
            if not col_name or col_name not in header:
                continue
            col_idx = header.index(col_name)
            for value, hex6 in palette.items():
                requests.append({"addConditionalFormatRule": {
                    "rule": {
                        "ranges": [{
                            "sheetId": sheet_id,
                            "startRowIndex": body_start,
                            "startColumnIndex": col_idx,
                            "endColumnIndex": col_idx + 1,
                        }],
                        "booleanRule": {
                            "condition": {
                                "type": "TEXT_EQ",
                                "values": [{"userEnteredValue": value}],
                            },
                            "format": {
                                "backgroundColorStyle": {"rgbColor": _hex_to_rgb(hex6)},
                            },
                        },
                    },
                    "index": 0,
                }})

        #  Data validation — quality dropdown (lenient: showWarning
        #  on bad input) and beam dropdown (STRICT: rejects anything
        #  not in ON/OFF, because Beam is a binary that must stay
        #  clean for downstream tooling).
        if fmt.quality_column and fmt.quality_column in header:
            q_idx = header.index(fmt.quality_column)
            requests.append({"setDataValidation": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": body_start,
                    "startColumnIndex": q_idx,
                    "endColumnIndex": q_idx + 1,
                },
                "rule": {
                    "condition": {
                        "type": "ONE_OF_LIST",
                        "values": [
                            {"userEnteredValue": v}
                            for v in QUALITY_PALETTE.keys()
                        ],
                    },
                    "showCustomUi": True,
                    "strict": False,
                },
            }})

        #  Body-wide centering (Phase D.8) — operators want every
        #  column centered so numbers + tags + firmware strings all
        #  align visually.  Applied via a single ``repeatCell`` over
        #  the data block; fields mask scopes the merge to alignment
        #  only so bold + chip colours from later passes stay intact.
        if nrows > body_start and ncols > 0:
            requests.append({"repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": body_start, "endRowIndex": nrows,
                    "startColumnIndex": 0, "endColumnIndex": ncols,
                },
                "cell": {"userEnteredFormat": {
                    "horizontalAlignment": "CENTER",
                    "verticalAlignment": "MIDDLE",
                }},
                "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment)",
            }})
            #  Phase D.11: per-column left-alignment override for the
            #  long-text columns (Notes).  Issued after the body-wide
            #  centering so it wins; the ``horizontalAlignment`` fields
            #  mask scopes the merge so chip colours / bold stay put.
            for col_name in fmt.left_aligned_columns:
                if col_name not in header:
                    continue
                col_idx = header.index(col_name)
                requests.append({"repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": body_start, "endRowIndex": nrows,
                        "startColumnIndex": col_idx, "endColumnIndex": col_idx + 1,
                    },
                    "cell": {"userEnteredFormat": {
                        "horizontalAlignment": "LEFT",
                    }},
                    "fields": "userEnteredFormat.horizontalAlignment",
                }})

        #  Tag-style each palette-bound column: bold + centered text
        #  on the body range so the coloured fill reads as a chip.
        #  Identical look across xlsx + live Sheet because both
        #  renderers consult ``_tag_columns_for(fmt)``.
        for col_name in _tag_columns_for(fmt):
            if col_name not in header:
                continue
            col_idx = header.index(col_name)
            requests.append({"repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": body_start,
                    "startColumnIndex": col_idx,
                    "endColumnIndex": col_idx + 1,
                },
                "cell": {"userEnteredFormat": {
                    "textFormat": {"bold": True},
                    "horizontalAlignment": "CENTER",
                    "verticalAlignment": "MIDDLE",
                }},
                "fields": "userEnteredFormat(textFormat.bold,horizontalAlignment,verticalAlignment)",
            }})

        if fmt.beam_column and fmt.beam_column in header:
            b_idx = header.index(fmt.beam_column)
            requests.append({"setDataValidation": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": body_start,
                    "startColumnIndex": b_idx,
                    "endColumnIndex": b_idx + 1,
                },
                "rule": {
                    "condition": {
                        "type": "ONE_OF_LIST",
                        "values": [
                            {"userEnteredValue": v}
                            for v in BEAM_PALETTE.keys()
                        ],
                    },
                    "showCustomUi": True,
                    "strict": True,   # binary — no free-form input allowed
                },
            }})

        #  Polarity dropdown — strict like Beam (Phase D.10).
        if fmt.polarity_column and fmt.polarity_column in header:
            p_idx = header.index(fmt.polarity_column)
            requests.append({"setDataValidation": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": body_start,
                    "startColumnIndex": p_idx,
                    "endColumnIndex": p_idx + 1,
                },
                "rule": {
                    "condition": {
                        "type": "ONE_OF_LIST",
                        "values": [
                            {"userEnteredValue": v}
                            for v in POLARITY_PALETTE.keys()
                        ],
                    },
                    "showCustomUi": True,
                    "strict": True,
                },
            }})

        #  Numeric format overrides (Phase D.10) — Temp X.X, V_bias X.XX.
        #  Applied as one repeatCell per column over the body range so
        #  the trailing-zero pattern persists across operator edits.
        for col_name, pattern in fmt.numeric_formats.items():
            if col_name not in header:
                continue
            n_idx = header.index(col_name)
            requests.append({"repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": body_start,
                    "endRowIndex": nrows,
                    "startColumnIndex": n_idx,
                    "endColumnIndex": n_idx + 1,
                },
                "cell": {"userEnteredFormat": {
                    "numberFormat": {"type": "NUMBER", "pattern": pattern},
                }},
                "fields": "userEnteredFormat.numberFormat",
            }})

    return requests


__all__ = [
    "AUDIT_SOURCE_PALETTE",
    "BEAM_PALETTE",
    "FORMATS",
    "JOB_STATE_PALETTE",
    "QUALITY_PALETTE",
    "SheetFormat",
    "TAB_COLORS",
    "base_name",
    "sheets_format_requests",
    "to_xlsx",
]
