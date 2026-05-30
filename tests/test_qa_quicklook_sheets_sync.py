"""Unit tests for ``qa_quicklook.sheets_sync``.

Pure-stdlib — no Qt, no network.  Exercises every deterministic
piece of the cross-shifter-sync feature: config loading, snapshot
building from a synthetic ``run-lists/`` tree, worksheet rendering,
the cell-diff reverse-merge logic, and snapshot persistence.

The Google adapter (``_sheets_adapter``) is tested separately because
its only non-trivial logic — HTTP error translation — needs a small
stub for the googleapiclient HttpError type.  Push/pull themselves
are thin wrappers and would require the real library to test
meaningfully, so they're covered by the manual GCP setup runbook
rather than by unit tests.

Run::

    .venv/bin/python -m unittest tests.test_qa_quicklook_sheets_sync
"""

from __future__ import annotations

import json
import os
import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

from qa_quicklook import sheets_sync  # noqa: E402


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


_DB_TOML = """\
[runs."20251109-135431"]
beam_energy = 180
beam_status = "on"
temperature = -28
v_bias = 45.00
n_spills = 10
quality = "warning"
radiators = [
  { type = "Aerogel", refindex = 1.021, tag = "AG22-J001", depth = 2 }
]
notes = "first run"

[runs."20251109-140610"]
n_spills = 50
notes = "second run"

[runs."20251110-115618"]
v_bias = 52.00
notes = "third"
quality = "good"
"""

_RUNLISTS_TOML = """\
[runlists.alpha]
runs = ["20251109-135431", "20251109-140610"]

[runlists.beta]
runs = ["20251110-115618"]
"""

_AUDIT_TOML = """\
[[entry]]
at        = "2026-05-29T10:00:00"
source    = "dashboard"
run       = "20251110-115618"
field     = "quality"
old_value = ""
new_value = "good"

[[entry]]
at        = "2026-05-29T10:01:00"
source    = "legacy"
run       = "20251109-135431"
field     = "beam_energy"
old_value = ""
new_value = 180
"""


class _FixtureRepo:
    """Sets up a tmp ``run-lists/`` tree mirroring the production layout."""

    def __init__(self) -> None:
        self._td = tempfile.TemporaryDirectory()
        self.root = Path(self._td.name)
        rl = self.root / "run-lists"
        rl.mkdir()
        (rl / "2025.database.toml").write_text(_DB_TOML)
        (rl / "2025.runlists.toml").write_text(_RUNLISTS_TOML)
        (rl / "2025.database.audit.toml").write_text(_AUDIT_TOML)

    def cleanup(self) -> None:
        self._td.cleanup()


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------


class TestSheetsConfig(unittest.TestCase):
    def setUp(self) -> None:
        self._td = tempfile.TemporaryDirectory()
        self.cfg_path = Path(self._td.name) / "qa_quicklook.toml"

    def tearDown(self) -> None:
        self._td.cleanup()

    def test_missing_file_defaults_disabled(self) -> None:
        cfg = sheets_sync.load_config(self.cfg_path)
        self.assertFalse(cfg.enabled)
        self.assertFalse(cfg.is_configured)
        self.assertIn("enabled = false", cfg.disabled_reason())

    def test_section_missing_defaults_disabled(self) -> None:
        self.cfg_path.write_text("[ui]\ntheme = \"dark\"\n")
        cfg = sheets_sync.load_config(self.cfg_path)
        self.assertFalse(cfg.enabled)
        self.assertFalse(cfg.is_configured)

    def test_enabled_but_unconfigured(self) -> None:
        # enabled=true but no spreadsheet_id → disabled_reason names
        # the missing piece so the operator knows what to fix.
        self.cfg_path.write_text(
            "[sheets_sync]\nenabled = true\n"
        )
        cfg = sheets_sync.load_config(self.cfg_path)
        self.assertTrue(cfg.enabled)
        self.assertFalse(cfg.is_configured)
        self.assertIn("spreadsheet_id", cfg.disabled_reason())

    def test_service_account_path_expands_tilde(self) -> None:
        self.cfg_path.write_text(
            "[sheets_sync]\n"
            "enabled = true\n"
            "spreadsheet_id = \"abc123\"\n"
            "service_account = \"~/sa.json\"\n"
        )
        cfg = sheets_sync.load_config(self.cfg_path)
        self.assertTrue(cfg.service_account_path.is_absolute())
        self.assertFalse(str(cfg.service_account_path).startswith("~"))

    def test_push_interval_floor(self) -> None:
        # A misconfigured 1-second interval should not slip through —
        # we want a floor so the Sheets API quota stays comfortable.
        self.cfg_path.write_text(
            "[sheets_sync]\nenabled = true\npush_interval_s = 1\n"
        )
        cfg = sheets_sync.load_config(self.cfg_path)
        self.assertGreaterEqual(cfg.push_interval_s, 5)

    def test_is_configured_requires_real_file(self) -> None:
        td = Path(self._td.name)
        sa_path = td / "sa.json"
        self.cfg_path.write_text(
            "[sheets_sync]\n"
            "enabled = true\n"
            "spreadsheet_id = \"abc\"\n"
            f"service_account = \"{sa_path}\"\n"
        )
        cfg = sheets_sync.load_config(self.cfg_path)
        self.assertFalse(cfg.is_configured)  # SA file doesn't exist yet
        sa_path.write_text("{}")              # whatever; presence is what matters
        cfg2 = sheets_sync.load_config(self.cfg_path)
        self.assertTrue(cfg2.is_configured)


class TestResolveOperatorTag(unittest.TestCase):
    def test_configured_wins_verbatim(self) -> None:
        self.assertEqual(
            sheets_sync.resolve_operator_tag("Mario, shifter 11/29"),
            "Mario, shifter 11/29",
        )

    def test_blank_falls_back_to_user_at_host(self) -> None:
        tag = sheets_sync.resolve_operator_tag("")
        self.assertIn("@", tag)
        self.assertNotEqual(tag, "@")


class TestAutoYear(unittest.TestCase):
    def test_picks_newest_database(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            rl = root / "run-lists"
            rl.mkdir()
            (rl / "2024.database.toml").write_text("")
            (rl / "2025.database.toml").write_text("")
            (rl / "2026.database.toml").write_text("")
            (rl / "2024.runlists.toml").write_text("")   # noise
            self.assertEqual(sheets_sync.auto_year(root), "2026")

    def test_missing_dir_returns_none(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            self.assertIsNone(sheets_sync.auto_year(Path(td)))


# ---------------------------------------------------------------------------
# Source-file readers
# ---------------------------------------------------------------------------


class TestSourceFileReaders(unittest.TestCase):
    def setUp(self) -> None:
        self.fx = _FixtureRepo()

    def tearDown(self) -> None:
        self.fx.cleanup()

    def test_read_runlists_meta(self) -> None:
        from qa_quicklook import rundb
        out = rundb.read_runlists_meta(
            self.fx.root / "run-lists" / "2025.runlists.toml"
        )
        self.assertEqual(set(out.keys()), {"alpha", "beta"})
        # Phase B shape: {name: {"runs": [...], "campaign": "..."}}
        self.assertEqual(out["alpha"]["runs"][0], "20251109-135431")
        self.assertEqual(out["alpha"]["campaign"], "")   # legacy file, no key

    def test_read_audit_tail_newest_first(self) -> None:
        out = sheets_sync._read_audit_tail(
            self.fx.root / "run-lists" / "2025.database.audit.toml",
            tail_n=10,
        )
        # The audit file was written oldest→newest; reader reverses
        # so the head row is the most recent edit.
        self.assertEqual(out[0]["at"], "2026-05-29T10:01:00")
        self.assertEqual(len(out), 2)

    def test_read_audit_tail_bounded(self) -> None:
        out = sheets_sync._read_audit_tail(
            self.fx.root / "run-lists" / "2025.database.audit.toml",
            tail_n=1,
        )
        self.assertEqual(len(out), 1)


# ---------------------------------------------------------------------------
# Snapshot
# ---------------------------------------------------------------------------


class TestBuildSnapshot(unittest.TestCase):
    """Multi-year Snapshot (Phase B).  Fixture has just 2025."""

    def setUp(self) -> None:
        self.fx = _FixtureRepo()

    def tearDown(self) -> None:
        self.fx.cleanup()

    def test_merged_view_includes_inherited_fields(self) -> None:
        snap = sheets_sync.build_snapshot(self.fx.root)
        self.assertEqual(snap.years, ["2025"])
        self.assertEqual(snap.year, "2025")           # back-compat property
        runs_2025 = snap.runs_by_year["2025"]
        self.assertEqual(len(runs_2025), 3)
        #  Forward inheritance: run 2 inherits beam_energy from run 1.
        run2 = next(r for r in runs_2025 if r["run_id"] == "20251109-140610")
        self.assertEqual(run2["beam_energy"], 180)
        self.assertEqual(run2["n_spills"], 50)

    def test_runlists_picked_up_per_year(self) -> None:
        snap = sheets_sync.build_snapshot(self.fx.root)
        self.assertIn("2025", snap.runlists_by_year)
        runlists_2025 = snap.runlists_by_year["2025"]
        self.assertIn("alpha", runlists_2025)
        # Phase B shape: per-runlist dict with runs + campaign.
        self.assertEqual(len(runlists_2025["alpha"]["runs"]), 2)
        self.assertEqual(runlists_2025["alpha"]["campaign"], "")

    def test_audit_tail_bounded(self) -> None:
        snap = sheets_sync.build_snapshot(self.fx.root, audit_tail_n=1)
        self.assertEqual(len(snap.audit_tail), 1)

    def test_radiators_catalog_built(self) -> None:
        # Phase B v2: catalog de-dupes per (type, refindex, tag, depth, side).
        # Fixture has one unique radiator inherited onto all 3 runs
        # → one catalog entry, three reference-map rows.  Phase D.16:
        # one contiguous range covers all three runs.
        snap = sheets_sync.build_snapshot(self.fx.root)
        self.assertEqual(len(snap.radiators_catalog), 1)
        entry = snap.radiators_catalog[0]
        self.assertEqual(entry.id, 1)
        self.assertEqual(entry.type, "Aerogel")
        self.assertEqual(entry.tag, "AG22-J001")
        self.assertEqual(entry.n_runs, 3)
        # Contiguous use → one range from first to last fixture run.
        self.assertEqual(entry.runs_ranges,
                         [("20251109-135431", "20251110-115618")])
        # All three runs reference catalog id 1.
        self.assertEqual(len(snap.radiator_ids_by_run), 3)
        for run_id in ("20251109-135431", "20251109-140610", "20251110-115618"):
            self.assertEqual(snap.radiator_ids_by_run[("2025", run_id)], [1])

    def test_radiators_catalog_splits_on_gap(self) -> None:
        # Phase D.16 — when a radiator is swapped out and back, the
        # catalog records two ranges, not one misleading "first to last"
        # span.  Synthesise a tiny snapshot directly.
        snap = sheets_sync.Snapshot(
            years=["2025"],
            runs_by_year={"2025": [
                {"run_id": "20251101-000001",
                 "radiators": [{"type": "Aerogel", "tag": "X",
                                "refindex": 1.02, "depth": 2}]},
                {"run_id": "20251101-000002",
                 "radiators": [{"type": "Aerogel", "tag": "X",
                                "refindex": 1.02, "depth": 2}]},
                {"run_id": "20251101-000003",
                 "radiators": [{"type": "Aerogel", "tag": "Y",
                                "refindex": 1.03, "depth": 2}]},
                {"run_id": "20251101-000004",
                 "radiators": [{"type": "Aerogel", "tag": "X",
                                "refindex": 1.02, "depth": 2}]},
            ]},
        )
        catalog, _ref = sheets_sync._build_radiators_catalog(snap.runs_by_year)
        x = next(e for e in catalog if e.tag == "X")
        y = next(e for e in catalog if e.tag == "Y")
        # X: two ranges (1–2 and 4–4).
        self.assertEqual(x.runs_ranges, [
            ("20251101-000001", "20251101-000002"),
            ("20251101-000004", "20251101-000004"),
        ])
        self.assertEqual(x.n_runs, 3)
        # Y: one range, single run.
        self.assertEqual(y.runs_ranges, [
            ("20251101-000003", "20251101-000003"),
        ])

    def test_format_runs_ranges_collapses_single_run(self) -> None:
        # Single-run ranges render as a bare id (no ``..`` suffix).
        out = sheets_sync._format_runs_ranges([
            ("R1", "R1"),
            ("R2", "R5"),
        ])
        self.assertEqual(out, "R1, R2..R5")

    def test_no_year_returns_empty(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            snap = sheets_sync.build_snapshot(Path(td))
            self.assertEqual(snap.years, [])
            self.assertEqual(snap.runs_by_year, {})
            self.assertIsNone(snap.year)


# ---------------------------------------------------------------------------
# Worksheet rendering
# ---------------------------------------------------------------------------


class TestRenderWorksheets(unittest.TestCase):
    """Phase B layout: row 0 = banner, row 1 = header, row 2+ = data."""

    def setUp(self) -> None:
        self.fx = _FixtureRepo()
        self.snap = sheets_sync.build_snapshot(
            self.fx.root, last_push_at="2026-05-29T12:00:00",
        )
        self.ws = sheets_sync.render_worksheets(self.snap)

    def tearDown(self) -> None:
        self.fx.cleanup()

    def test_worksheet_set_matches_phase_b(self) -> None:
        self.assertEqual(
            set(self.ws.keys()),
            {"Runs (2025)", "Runlists", "Radiators", "Audit"},
        )

    def test_banner_row_carries_last_push(self) -> None:
        for name, rows in self.ws.items():
            self.assertGreaterEqual(len(rows), 1, f"{name} has no rows")
            banner = rows[0][0]
            self.assertTrue(
                banner.startswith("Last push:"),
                f"{name} row 0 should be banner; got {banner!r}",
            )
            self.assertIn("2026-05-29T12:00:00", banner)

    def test_qa_results_tab_dropped(self) -> None:
        # Phase B v2: qa_results moved off its own tab — those values
        # belong as columns on runs (YYYY), not as a separate entity.
        # The xlsx + Sheet should not carry a qa_results tab today;
        # when standard_results.toml gets populated we'll pivot it
        # into qa_<quantity> columns on runs.
        for name in self.ws.keys():
            self.assertFalse(
                name.startswith("qa_results"),
                f"qa_results tab leaked back into render: {name}",
            )

    def test_runs_header_starts_with_canonical_fields(self) -> None:
        # Phase D.5: row 0 = banner, row 1 = group band, row 2 = header.
        header = self.ws["Runs (2025)"][2]
        # Phase D.1: header carries display names now.
        self.assertEqual(header[0], "Run ID")
        idx = {col: i for i, col in enumerate(header)}
        # Phase D.5 ordering: Beam group (Status + Energy) before
        # Detector group (Temp + V_bias).  Phase D.7: dropped group-
        # redundant prefixes.  Phase D.8: Temperature → Temp.
        self.assertLess(idx["Status"], idx["Energy (GeV)"])
        self.assertLess(idx["Energy (GeV)"], idx["Temp (°C)"])
        self.assertLess(idx["Temp (°C)"], idx["V_bias (V)"])

    def test_runs_excludes_radiators_column(self) -> None:
        # Phase B: radiators moved to its own worksheet.  Runs sheet
        # must NOT carry the column anymore.
        header = self.ws["Runs (2025)"][1]
        self.assertNotIn("Radiators", header)

    def test_runs_cell_preserves_scalar_types(self) -> None:
        # Phase D.5 layout: row 2 header, row 3 first data.
        # Phase D.7: column names dropped group-redundant prefix.
        header = self.ws["Runs (2025)"][2]
        first_data = self.ws["Runs (2025)"][3]   # 20251109-135431
        cells = dict(zip(header, first_data))
        self.assertIsInstance(cells["Energy (GeV)"], int)
        self.assertIsInstance(cells["V_bias (V)"], float)
        # Phase D.2: beam_status normalised "on" → "ON" on display.
        self.assertEqual(cells["Status"], "ON")
        # Phase D.8: Temperature → Temp.
        self.assertEqual(cells["Temp (°C)"], -28)

    def test_radiators_catalog_dedupes(self) -> None:
        # Phase B v2: radiators is now a de-duplicated catalog.  Fixture
        # has one unique radiator inherited across all 3 runs, so
        # there's exactly one catalog row regardless of usage count.
        header = self.ws["Radiators"][1]
        # Phase D.16: First/Last collapsed into a single "Runs ranges" cell.
        self.assertEqual(
            header,
            ["ID", "Type", "Refractive index", "Tag", "Depth (cm)",
             "Side", "N runs", "Runs ranges"],
        )
        # Data rows: banner + header + N catalog entries.
        data = self.ws["Radiators"][2:]
        self.assertEqual(len(data), 1)
        row = data[0]
        self.assertEqual(row[0], 1)                     # catalog id
        self.assertEqual(row[1], "Aerogel")             # type
        self.assertEqual(row[3], "AG22-J001")           # tag
        self.assertEqual(row[6], 3)                     # n_runs
        # Contiguous use → one range "first..last".
        self.assertEqual(row[7],
                         "20251109-135431..20251110-115618")

    def test_runs_tab_carries_radiator_ids(self) -> None:
        # Phase D.5: row 2 header, row 3 first data.
        header = self.ws["Runs (2025)"][2]
        self.assertIn("Radiator IDs", header)
        idx = header.index("Radiator IDs")
        first_data = self.ws["Runs (2025)"][3]
        # Fixture's three runs all share one radiator → "1".
        self.assertEqual(first_data[idx], "1")

    def test_runlists_vertical_layout(self) -> None:
        # Phase D.4: single cross-year Runlists tab.  Rows:
        #   0 banner / 1 Name / 2 Year / 3 Campaign / 4 N runs / 5+ run ids.
        rows = self.ws["Runlists"]
        self.assertEqual(rows[1][0], "Name")
        # Both fixture runlists (year 2025) should appear as column headers.
        self.assertIn("alpha", rows[1])
        self.assertIn("beta", rows[1])
        self.assertEqual(rows[2][0], "Year")
        self.assertEqual(rows[3][0], "Campaign")
        self.assertEqual(rows[4][0], "N runs")
        alpha_col_idx = rows[1].index("alpha")
        self.assertEqual(rows[2][alpha_col_idx], "2025")   # fixture year
        self.assertEqual(rows[4][alpha_col_idx], 2)        # alpha has 2 runs

    def test_audit_header(self) -> None:
        # Phase B: header on row 1, banner on row 0.  Phase D.1: display names.
        self.assertEqual(
            self.ws["Audit"][1],
            ["At", "Source", "Run", "Field", "Old value", "New value"],
        )

    def test_no_meta_no_jobs(self) -> None:
        # Phase B: meta + jobs sheets dropped.
        self.assertNotIn("meta", self.ws)
        self.assertNotIn("jobs", self.ws)


# ---------------------------------------------------------------------------
# Cell index + reverse-merge logic
# ---------------------------------------------------------------------------


class TestRunsCellIndex(unittest.TestCase):
    def test_skips_run_id_and_non_scalar_columns(self) -> None:
        rendered = [
            ["run_id", "quality", "radiators", "beam_energy"],
            ["R1", "good", "[]", 180],
            ["R2", "bad", "[]", 120],
        ]
        idx = sheets_sync.runs_cell_index(rendered)
        # run_id column itself is the key, not a value:
        self.assertNotIn(("R1", "run_id"), idx)
        # radiators is non-scalar so reverse-merge skips it:
        self.assertNotIn(("R1", "radiators"), idx)
        self.assertEqual(idx[("R1", "quality")], "good")
        self.assertEqual(idx[("R2", "beam_energy")], 120)

    def test_empty_input(self) -> None:
        self.assertEqual(sheets_sync.runs_cell_index([]), {})

    def test_missing_run_id_column(self) -> None:
        # Defensive: header has no run_id column at all.
        rendered = [["foo", "bar"], ["1", "2"]]
        self.assertEqual(sheets_sync.runs_cell_index(rendered), {})


class TestDetectReverseEdits(unittest.TestCase):
    """Three buckets the diff can put a cell in: apply / skip-in-sync / no-op."""

    def _ws(self, *rows):
        return [["run_id", "quality", "notes"], *rows]

    def test_apply_real_sheet_edit(self) -> None:
        sheet = self._ws(["R1", "bad", "edited on Sheet"])
        local = self._ws(["R1", "good", "untouched"])
        snap  = self._ws(["R1", "good", "untouched"])
        edits = sheets_sync.detect_reverse_edits(sheet, local, snap)
        applies = [e for e in edits if e.action == "apply"]
        self.assertEqual(len(applies), 2)   # quality + notes
        keys = {(e.run_id, e.field) for e in applies}
        self.assertIn(("R1", "quality"), keys)
        self.assertIn(("R1", "notes"), keys)

    def test_skip_when_already_in_sync(self) -> None:
        # Sheet edit beat the snapshot but the dashboard already
        # landed the same value — replaying would be a no-op edit
        # that still pollutes the audit log.  Diff classifies as
        # skip_already_in_sync so the caller drops it.
        sheet = self._ws(["R1", "bad", "x"])
        local = self._ws(["R1", "bad", "x"])
        snap  = self._ws(["R1", "good", "y"])
        edits = sheets_sync.detect_reverse_edits(sheet, local, snap)
        for e in edits:
            self.assertEqual(e.action, "skip_already_in_sync")

    def test_no_change_emits_nothing(self) -> None:
        # Identical Sheet vs snapshot → empty diff.
        rows = ["R1", "good", "x"]
        edits = sheets_sync.detect_reverse_edits(
            self._ws(rows), self._ws(rows), self._ws(rows),
        )
        self.assertEqual(edits, [])

    def test_type_tolerant_compare(self) -> None:
        # Sheets returns 180 as "180" on read; the diff should not
        # treat that as a Sheet edit.
        sheet = self._ws(["R1", "180", ""])
        local = self._ws(["R1", 180, ""])
        snap  = self._ws(["R1", 180, ""])
        edits = sheets_sync.detect_reverse_edits(sheet, local, snap)
        # Quality column changed from 180→"180" but they canonicalise
        # to the same string, so no edit.
        applies = [e for e in edits if e.action == "apply"]
        self.assertEqual(applies, [])


class TestCanonicalCompare(unittest.TestCase):
    def test_bool_renders_uppercase(self) -> None:
        self.assertEqual(sheets_sync._canonical(True), "TRUE")
        self.assertEqual(sheets_sync._canonical(False), "FALSE")

    def test_float_keeps_precision(self) -> None:
        self.assertEqual(sheets_sync._canonical(45.0), "45.0")

    def test_strip_strings(self) -> None:
        self.assertEqual(sheets_sync._canonical("  hi  "), "hi")

    def test_none_collapses_to_empty(self) -> None:
        self.assertTrue(sheets_sync._cells_equal(None, ""))


class TestCoerceSheetValue(unittest.TestCase):
    def test_int_round_trip(self) -> None:
        self.assertEqual(sheets_sync._coerce_sheet_value("180", 0), 180)

    def test_float_round_trip(self) -> None:
        self.assertEqual(sheets_sync._coerce_sheet_value("45.5", 0.0), 45.5)

    def test_bool_round_trip(self) -> None:
        self.assertIs(sheets_sync._coerce_sheet_value("TRUE", True), True)
        self.assertIs(sheets_sync._coerce_sheet_value("no", False), False)

    def test_string_fallback_on_failed_cast(self) -> None:
        # "45 V" can't be parsed as int — we keep the literal so the
        # operator's text round-trips visibly through the next push.
        self.assertEqual(sheets_sync._coerce_sheet_value("45 V", 0), "45 V")


class TestApplyReverseEdits(unittest.TestCase):
    """Verify replays land via ``rundb.update_run_field`` with source='sheet'.

    Phase B: signature changed from ``(edits, db_path)`` to
    ``(edits, repo_root)``; each edit picks its database from the
    year prefix of its run_id (``20251109-…`` → ``2025.database.toml``).
    """

    def setUp(self) -> None:
        self.fx = _FixtureRepo()

    def tearDown(self) -> None:
        self.fx.cleanup()

    def test_applies_and_skips(self) -> None:
        edits = [
            sheets_sync.SheetEdit(
                run_id="20251109-135431", field="quality",
                snapshot_value="good", local_value="good",
                sheet_value="bad", action="apply",
            ),
            sheets_sync.SheetEdit(
                run_id="20251109-140610", field="notes",
                snapshot_value="x", local_value="y",
                sheet_value="y", action="skip_already_in_sync",
            ),
        ]
        with mock.patch(
            "qa_quicklook.rundb.update_run_field"
        ) as patched:
            applied, skipped = sheets_sync.apply_reverse_edits(
                edits, self.fx.root,
            )
        self.assertEqual(applied, 1)
        self.assertEqual(skipped, 1)
        patched.assert_called_once()
        kwargs = patched.call_args.kwargs
        self.assertEqual(kwargs["source"], "sheet")
        # Args: (db_path, run, field, value); db_path derived from year.
        args = patched.call_args.args
        self.assertEqual(args[0].name, "2025.database.toml")
        self.assertEqual(args[1], "20251109-135431")
        self.assertEqual(args[2], "quality")
        self.assertEqual(args[3], "bad")

    def test_skips_runs_without_year_prefix(self) -> None:
        # Defensive: a sheet edit with a malformed run id (e.g. someone
        # pasted "??? unknown") should skip without crashing.
        edits = [
            sheets_sync.SheetEdit(
                run_id="??? unknown", field="quality",
                snapshot_value="", local_value="",
                sheet_value="bad", action="apply",
            ),
        ]
        with mock.patch(
            "qa_quicklook.rundb.update_run_field"
        ) as patched:
            applied, skipped = sheets_sync.apply_reverse_edits(
                edits, self.fx.root,
            )
        self.assertEqual(applied, 0)
        self.assertEqual(skipped, 1)
        patched.assert_not_called()

    def test_skips_runs_with_strict_pattern_typo(self) -> None:
        # Phase D.13: the strict run-id regex catches typos that
        # would otherwise auto-append a phantom row.  Example:
        # someone types "20260601" (date only, no time) into the
        # Sheet's Run ID column — we skip rather than insert.
        for bad in ("20260601", "20260601-10", "2026/06/01-103022",
                    "  20260601-103022  whoops"):
            edits = [
                sheets_sync.SheetEdit(
                    run_id=bad, field="quality",
                    snapshot_value="", local_value="",
                    sheet_value="bad", action="apply",
                ),
            ]
            with mock.patch(
                "qa_quicklook.rundb.update_run_field"
            ) as patched_update, mock.patch(
                "qa_quicklook.rundb.append_runs"
            ) as patched_append:
                applied, skipped = sheets_sync.apply_reverse_edits(
                    edits, self.fx.root,
                )
            self.assertEqual(applied, 0,
                             f"{bad!r} should be skipped, was applied")
            self.assertEqual(skipped, 1, f"{bad!r}")
            patched_update.assert_not_called()
            patched_append.assert_not_called()

    def test_auto_appends_new_run_id(self) -> None:
        # Phase D.13: a Sheet edit on a run id that doesn't yet exist
        # in the local TOML triggers an append_runs + update_run_field
        # sequence — no more silent drops.
        new_id = "20251231-235959"   # fits the strict pattern, 2025 year
        edits = [
            sheets_sync.SheetEdit(
                run_id=new_id, field="notes",
                snapshot_value="", local_value="",
                sheet_value="spontaneously added from Sheet",
                action="apply",
            ),
        ]
        with mock.patch(
            "qa_quicklook.rundb.update_run_field"
        ) as patched_update, mock.patch(
            "qa_quicklook.rundb.append_runs"
        ) as patched_append:
            applied, skipped = sheets_sync.apply_reverse_edits(
                edits, self.fx.root,
            )
        self.assertEqual(applied, 1)
        self.assertEqual(skipped, 0)
        # append_runs called once for the new id BEFORE the field edit.
        patched_append.assert_called_once()
        args, _ = patched_append.call_args
        self.assertEqual(args[0].name, "2025.database.toml")
        self.assertEqual(args[1], [new_id])
        # Then the field edit lands.
        patched_update.assert_called_once()
        u_args = patched_update.call_args.args
        self.assertEqual(u_args[1], new_id)
        self.assertEqual(u_args[2], "notes")


# ---------------------------------------------------------------------------
# Snapshot persistence
# ---------------------------------------------------------------------------


class TestSnapshotPersistence(unittest.TestCase):
    def setUp(self) -> None:
        self._td = tempfile.TemporaryDirectory()
        self.path = Path(self._td.name) / "snap.json"

    def tearDown(self) -> None:
        self._td.cleanup()

    def test_round_trip(self) -> None:
        payload = {
            "runs": [["run_id", "quality"], ["R1", "good"]],
            "meta": [["key", "value"], ["year", "2025"]],
        }
        sheets_sync.save_last_pushed(payload, path=self.path)
        loaded = sheets_sync.load_last_pushed(path=self.path)
        self.assertEqual(loaded["runs"], payload["runs"])
        self.assertEqual(loaded["meta"], payload["meta"])

    def test_missing_file_returns_empty(self) -> None:
        self.assertEqual(sheets_sync.load_last_pushed(path=self.path), {})

    def test_corrupt_file_returns_empty(self) -> None:
        self.path.write_text("{not json")
        self.assertEqual(sheets_sync.load_last_pushed(path=self.path), {})

    def test_creates_parent_dir(self) -> None:
        nested = Path(self._td.name) / "a" / "b" / "snap.json"
        sheets_sync.save_last_pushed({"runs": []}, path=nested)
        self.assertTrue(nested.is_file())


# ---------------------------------------------------------------------------
# End-to-end dry-run sync
# ---------------------------------------------------------------------------


class TestSyncOnceDryRun(unittest.TestCase):
    """``sync_once(dry_run=True)`` exercises the full pipeline minus network."""

    def setUp(self) -> None:
        self.fx = _FixtureRepo()
        self.snap_path = self.fx.root / "snap.json"

    def tearDown(self) -> None:
        self.fx.cleanup()

    def test_dry_run_renders_without_touching_google(self) -> None:
        cfg = sheets_sync.SheetsConfig(enabled=False)
        result = sheets_sync.sync_once(
            cfg, self.fx.root,
            dry_run=True, snapshot_path=self.snap_path,
        )
        self.assertTrue(result.dry_run)
        self.assertEqual(result.year, "2025")
        # Phase B worksheet inventory.
        self.assertEqual(
            set(result.rendered.keys()),
            {"Runs (2025)", "Runlists", "Radiators", "Audit"},
        )
        # Snapshot is NOT written in dry-run — that would poison the
        # next real push's reverse-merge baseline.
        self.assertFalse(self.snap_path.exists())


# ---------------------------------------------------------------------------
# xlsx formatter — only runs when openpyxl is importable.  These tests
# are the local-preview path's correctness check; the live Sheets API
# formatting uses the same SheetFormat specs.
# ---------------------------------------------------------------------------


try:
    import openpyxl  # noqa: F401
    _HAVE_OPENPYXL = True
except ImportError:
    _HAVE_OPENPYXL = False


@unittest.skipUnless(_HAVE_OPENPYXL, "openpyxl not installed")
class TestSheetsFormat(unittest.TestCase):
    """Phase B layout: row 1 = banner, row 2 = header, rows 3+ = data."""

    def setUp(self) -> None:
        self.fx = _FixtureRepo()
        self.snap = sheets_sync.build_snapshot(
            self.fx.root, last_push_at="2026-05-29T12:00:00",
        )
        self.rendered = sheets_sync.render_worksheets(self.snap)
        self._td = tempfile.TemporaryDirectory()
        self.out = Path(self._td.name) / "preview.xlsx"

    def tearDown(self) -> None:
        self._td.cleanup()
        self.fx.cleanup()

    def test_xlsx_round_trip(self) -> None:
        from qa_quicklook import _sheets_format
        written = _sheets_format.to_xlsx(self.rendered, self.out)
        self.assertEqual(written, self.out)
        self.assertTrue(self.out.is_file())
        from openpyxl import load_workbook
        wb = load_workbook(self.out)
        # Phase B inventory.
        self.assertEqual(
            set(wb.sheetnames),
            {"Runs (2025)", "Runlists", "Radiators", "Audit"},
        )

    def test_header_row_is_styled(self) -> None:
        from qa_quicklook import _sheets_format
        _sheets_format.to_xlsx(self.rendered, self.out)
        from openpyxl import load_workbook
        wb = load_workbook(self.out)
        ws = wb["Runs (2025)"]
        # Phase D.5: A1 banner / A2 group band / A3 column header.
        self.assertTrue(ws["A1"].value.startswith("Last push:"))
        self.assertEqual(ws["A3"].value, "Run ID")
        self.assertTrue(ws["A3"].font.bold)
        self.assertEqual(
            ws["A3"].fill.fgColor.rgb.upper()[-6:],
            _sheets_format.HEADER_BG_HEX,
        )

    def test_quality_column_color_coded(self) -> None:
        from qa_quicklook import _sheets_format
        _sheets_format.to_xlsx(self.rendered, self.out)
        from openpyxl import load_workbook
        wb = load_workbook(self.out)
        ws = wb["Runs (2025)"]
        # Phase D.5: row 3 is the column header (after banner + group band).
        header = [c.value for c in ws[3]]
        qcol = header.index("Quality") + 1
        # First data row is row 4.  Fixture has quality="warning".
        cell = ws.cell(row=4, column=qcol)
        self.assertEqual(cell.value, "warning")
        self.assertEqual(
            cell.fill.fgColor.rgb.upper()[-6:],
            _sheets_format.QUALITY_PALETTE["warning"],
        )

    def test_frozen_header_pane(self) -> None:
        from qa_quicklook import _sheets_format
        _sheets_format.to_xlsx(self.rendered, self.out)
        from openpyxl import load_workbook
        wb = load_workbook(self.out)
        # Runs (YYYY): frozen_rows=3, frozen_cols=3 (Phase D.9) →
        # freeze pane top-left at D4 (Run ID + N spills + Quality pinned).
        self.assertEqual(str(wb["Runs (2025)"].freeze_panes), "D4")
        # Runlists: frozen_rows=5 → A6 (banner + name + year + campaign + count).
        self.assertEqual(str(wb["Runlists"].freeze_panes), "A6")

    def test_no_qa_results_tab(self) -> None:
        # Phase B v2: dropped; values belong as columns on runs (YYYY).
        from qa_quicklook import _sheets_format
        _sheets_format.to_xlsx(self.rendered, self.out)
        from openpyxl import load_workbook
        wb = load_workbook(self.out)
        for name in wb.sheetnames:
            self.assertFalse(
                name.startswith("qa_results"),
                f"qa_results tab leaked into xlsx: {name}",
            )


# ---------------------------------------------------------------------------
# Phase A regression tests — the 2026-05-29 incident set
# ---------------------------------------------------------------------------


class TestTryFloat(unittest.TestCase):
    """``_try_float`` is the gate between numeric tolerance and false matches."""

    def test_int_and_float_pass_through(self) -> None:
        self.assertEqual(sheets_sync._try_float(45), 45.0)
        self.assertEqual(sheets_sync._try_float(45.5), 45.5)

    def test_string_numbers_parse(self) -> None:
        self.assertEqual(sheets_sync._try_float("45"), 45.0)
        self.assertEqual(sheets_sync._try_float("45.5"), 45.5)
        self.assertEqual(sheets_sync._try_float("  45.5  "), 45.5)

    def test_bool_returns_none(self) -> None:
        # ``True == 1.0`` is the Python quirk we don't want to leak —
        # ``quality`` cells should never numerically equate to bools.
        self.assertIsNone(sheets_sync._try_float(True))
        self.assertIsNone(sheets_sync._try_float(False))

    def test_garbage_returns_none(self) -> None:
        self.assertIsNone(sheets_sync._try_float("hi"))
        self.assertIsNone(sheets_sync._try_float("51,5"))   # the locale bug

    def test_none_and_empty_return_none(self) -> None:
        self.assertIsNone(sheets_sync._try_float(None))
        self.assertIsNone(sheets_sync._try_float(""))


class TestCellsEqualNumericTolerance(unittest.TestCase):
    """The fix for cross-type drift between Sheet pull + local TOML."""

    def test_int_equals_string_int(self) -> None:
        # 180 (int from local) vs "180" (str from canonical compare) —
        # already handled, kept here as a baseline.
        self.assertTrue(sheets_sync._cells_equal(180, "180"))

    def test_float_equals_int_when_value_matches(self) -> None:
        # 45.0 (local) vs 45 (Sheet read as int via UNFORMATTED_VALUE) —
        # this is the case that DID flip a row before A.1 shipped.
        self.assertTrue(sheets_sync._cells_equal(45.0, 45))

    def test_int_equals_float_string(self) -> None:
        # Defensive — if anything serialised the float as "45.0" str
        # somewhere, we still want the comparison to fire.
        self.assertTrue(sheets_sync._cells_equal(45, "45.0"))

    def test_locale_decimal_stays_unequal(self) -> None:
        # "51,5" (European decimal) is NOT numerically equivalent to
        # 51.5 in any locale-agnostic parser — we want the safety brake
        # to fire instead of silently mangling the database.
        self.assertFalse(sheets_sync._cells_equal(51.5, "51,5"))

    def test_distinct_floats_stay_unequal(self) -> None:
        self.assertFalse(sheets_sync._cells_equal(45.0, 45.5))

    def test_bool_does_not_equal_one(self) -> None:
        # Don't let True == 1 sneak through the numeric path.
        self.assertFalse(sheets_sync._cells_equal(True, 1))


class TestReverseMergeBrake(unittest.TestCase):
    """The safety brake on ``sync_once``."""

    def setUp(self) -> None:
        self.fx = _FixtureRepo()
        self._td = tempfile.TemporaryDirectory()
        self.snap_path = Path(self._td.name) / "snap.json"
        self.cfg = sheets_sync.SheetsConfig(enabled=False)   # dry-run only

    def tearDown(self) -> None:
        self._td.cleanup()
        self.fx.cleanup()

    def _fake_payload(self, n: int, *, with_banner: bool = True) -> tuple[list, list]:
        """Build (sheet_rows, local_rows) where each row's quality + notes differ.

        ``with_banner`` controls whether row 0 is a banner cell — matches
        the post-Phase-B rendered shape that ``runs_cell_index`` skips.
        """
        banner_row = [["Last push: 2026-05-29T00:00:00 — test"]] if with_banner else []
        header = ["run_id", "quality", "notes"]
        sheet = banner_row + [header] + [
            [f"20251109-{i:06d}", "fabricated", f"row-{i}"] for i in range(n)
        ]
        local = banner_row + [header] + [
            [f"20251109-{i:06d}", "good", "untouched"] for i in range(n)
        ]
        return sheet, local

    def test_brake_raises_when_threshold_exceeded(self) -> None:
        from unittest.mock import patch
        sheet_rows, local_rows = self._fake_payload(30)
        cfg = sheets_sync.SheetsConfig(
            enabled=True, spreadsheet_id="x",
            service_account="/dev/null",
        )
        with patch.object(sheets_sync, "_render_runs_year",
                          return_value=local_rows), \
             patch.object(sheets_sync, "check_sheet_integrity",
                          return_value=[]), \
             patch("qa_quicklook._sheets_adapter.pull_runs_worksheet",
                   return_value=sheet_rows):
            with self.assertRaises(sheets_sync.ReverseMergeBrake) as ctx:
                sheets_sync.sync_once(
                    cfg, self.fx.root,
                    snapshot_path=self.snap_path,
                    max_reverse_edits=10,
                )
            self.assertGreater(ctx.exception.edit_count, 10)
            self.assertEqual(ctx.exception.threshold, 10)
            db = self.fx.root / "run-lists" / "2025.database.toml"
            self.assertEqual(db.read_text(), db.read_text())

    def test_force_bypasses_brake(self) -> None:
        from unittest.mock import patch, MagicMock
        sheet_rows, local_rows = self._fake_payload(30)
        cfg = sheets_sync.SheetsConfig(
            enabled=True, spreadsheet_id="x",
            service_account="/dev/null",
        )
        applied_mock = MagicMock(return_value=(30, 0))
        with patch.object(sheets_sync, "_render_runs_year",
                          return_value=local_rows), \
             patch.object(sheets_sync, "check_sheet_integrity",
                          return_value=[]), \
             patch("qa_quicklook._sheets_adapter.pull_runs_worksheet",
                   return_value=sheet_rows), \
             patch.object(sheets_sync, "apply_reverse_edits", applied_mock), \
             patch("qa_quicklook._sheets_adapter.push_snapshot",
                   return_value=MagicMock(
                       last_push_at="2026-05-29T00:00:00",
                       updated_ranges=[], rows_changed=0,
                       worksheets_skipped=0,
                   )):
            sheets_sync.sync_once(
                cfg, self.fx.root,
                snapshot_path=self.snap_path,
                max_reverse_edits=10,
                force_reverse_merge=True,
            )
        self.assertEqual(applied_mock.call_count, 1)


class TestColLetter(unittest.TestCase):
    """A1-notation column letter generator."""

    def test_single_letter(self) -> None:
        from qa_quicklook import _sheets_adapter
        self.assertEqual(_sheets_adapter._col_letter(1), "A")
        self.assertEqual(_sheets_adapter._col_letter(26), "Z")

    def test_double_letter(self) -> None:
        from qa_quicklook import _sheets_adapter
        self.assertEqual(_sheets_adapter._col_letter(27), "AA")
        self.assertEqual(_sheets_adapter._col_letter(52), "AZ")
        self.assertEqual(_sheets_adapter._col_letter(702), "ZZ")

    def test_triple_letter(self) -> None:
        from qa_quicklook import _sheets_adapter
        self.assertEqual(_sheets_adapter._col_letter(703), "AAA")


class TestDiffAwarePush(unittest.TestCase):
    """Row-level diff produces the right batchUpdate payload."""

    def setUp(self) -> None:
        from unittest.mock import MagicMock
        self.svc = MagicMock()
        # Wire up the call chain so values().batchUpdate(...).execute()
        # returns a stub response.
        self.svc.values.return_value.batchUpdate.return_value.execute.return_value = {
            "totalUpdatedRows": 0,
        }
        self.svc.values.return_value.clear.return_value.execute.return_value = {}
        self.cfg = sheets_sync.SheetsConfig(
            enabled=True, spreadsheet_id="x",
            service_account="/dev/null",
        )

    def test_unchanged_row_skipped(self) -> None:
        from qa_quicklook import _sheets_adapter
        rows = [["a", "b"], ["1", "2"]]
        snapshot = [["a", "b"], ["1", "2"]]
        n, ranges = _sheets_adapter._diff_aware_update(
            self.svc, self.cfg, "runs", rows, snapshot,
        )
        self.assertEqual(n, 0)
        self.assertEqual(ranges, [])
        # batchUpdate must NOT have been called when nothing changed.
        self.svc.values.return_value.batchUpdate.assert_not_called()

    def test_single_row_changed(self) -> None:
        from qa_quicklook import _sheets_adapter
        rows = [["a", "b"], ["1", "2"], ["3", "FOUR"]]
        snapshot = [["a", "b"], ["1", "2"], ["3", "4"]]
        n, ranges = _sheets_adapter._diff_aware_update(
            self.svc, self.cfg, "runs", rows, snapshot,
        )
        self.assertEqual(n, 1)
        self.assertEqual(ranges, ["runs!A3:B3"])
        # batchUpdate was called once with the single-row payload.
        self.svc.values.return_value.batchUpdate.assert_called_once()
        body = self.svc.values.return_value.batchUpdate.call_args.kwargs["body"]
        self.assertEqual(len(body["data"]), 1)
        self.assertEqual(body["data"][0]["values"], [["3", "FOUR"]])

    def test_appended_row(self) -> None:
        from qa_quicklook import _sheets_adapter
        rows = [["a", "b"], ["1", "2"], ["3", "4"]]
        snapshot = [["a", "b"], ["1", "2"]]
        n, _ = _sheets_adapter._diff_aware_update(
            self.svc, self.cfg, "runs", rows, snapshot,
        )
        self.assertEqual(n, 1)   # one new row appended

    def test_removed_rows_cleared(self) -> None:
        from qa_quicklook import _sheets_adapter
        rows = [["a", "b"], ["1", "2"]]
        snapshot = [["a", "b"], ["1", "2"], ["3", "4"], ["5", "6"]]
        _sheets_adapter._diff_aware_update(
            self.svc, self.cfg, "runs", rows, snapshot,
        )
        # Trailing rows 3-4 should have been cleared with one call.
        clear_call = self.svc.values.return_value.clear.call_args
        self.assertIn("runs!A3:B4", clear_call.kwargs["range"])


class TestCheckSheetIntegrity(unittest.TestCase):
    """Phase D.14 — structural integrity check."""

    def test_no_issues_on_canonical_sheet(self) -> None:
        rendered = {
            "Runs (2025)": [
                ["Last push: …"],
                ["", "Beam", "", "Detector"],
                ["Run ID", "Status", "Particle", "Temp (°C)"],
                ["20251109-135431", "ON", "pions", "-28.0"],
            ],
        }
        sheet_rows = {"Runs (2025)": rendered["Runs (2025)"]}
        issues = sheets_sync.check_sheet_integrity(
            expected_titles=["Runs (2025)"],
            live_titles={"Runs (2025)"},
            local_rendered=rendered,
            sheet_rows_by_title=sheet_rows,
        )
        self.assertEqual(issues, [])

    def test_missing_tab(self) -> None:
        issues = sheets_sync.check_sheet_integrity(
            expected_titles=["Runs (2025)", "Runlists", "Audit"],
            live_titles={"Runs (2025)"},
            local_rendered={},
            sheet_rows_by_title={},
        )
        self.assertEqual(len(issues), 2)
        self.assertTrue(any("Runlists" in i for i in issues))
        self.assertTrue(any("Audit" in i for i in issues))

    def test_header_drift_pinpoints_first_divergence(self) -> None:
        # Operator renamed col 4 from "Temp (°C)" to "Temperature".
        local = [["banner"], ["", "Beam"],
                 ["Run ID", "Status", "Particle", "Temp (°C)"]]
        sheet = [["banner"], ["", "Beam"],
                 ["Run ID", "Status", "Particle", "Temperature"]]
        issues = sheets_sync.check_sheet_integrity(
            expected_titles=["Runs (2025)"],
            live_titles={"Runs (2025)"},
            local_rendered={"Runs (2025)": local},
            sheet_rows_by_title={"Runs (2025)": sheet},
        )
        self.assertEqual(len(issues), 1)
        self.assertIn("column 4", issues[0])
        self.assertIn("Temp (°C)", issues[0])
        self.assertIn("Temperature", issues[0])


if __name__ == "__main__":
    unittest.main()
